from math import nan
import PySimpleGUI as sg
import pandas as pd
import openpyxl
import numpy as np
from cgi import test
import tkinter as tk
from tkinter import font


class Application:

    def __init__(self, appName):
        self.jonesboroForecast_path = None
        self.choiceReport_path = None
        
        self.parts_changed = None
        self.parts_not_found = None
        self.JF_columns_to_transfer = None
        self.JF_column_names = None








        self.layout = [[sg.Button('Open Choice Report'), sg.Button('Open Jonesboro Forecast'), sg.Button('Do Data transfer')],
          [sg.Text(size=(40,1), key='-OUTPUT-')]]
        self.window = sg.Window(appName, self.layout)

    def read_data_files(self):
        try:
            data_CR = pd.read_excel(self.choiceReport_path)
            
            # workbook = openpyxl.load_workbook(self.choiceReport)
        except:
            self.window['-OUTPUT-'].update('Invalid file for Choice Report.')
            return
        try:
            data_JF = pd.read_excel(self.jonesboroForecast_path)
            
            # workbook = openpyxl.load_workbook(self.choiceReport)
        except:
            self.window['-OUTPUT-'].update('Invalid file for Jonesboro Forecast.')
            return
        
        # date = datetime.date.today()
        # sheet = workbook.active

        # cell = sheet.cell(row = 1, column = 4)
        # cell.value = ? #######



        ## Here we are constructing the data columns into CR and JF. Part number not included in this big array of arrays.

        CR_column_names = ["ICPROD","Total On Hand","On Order","Consigned","Required","SumOfWIP"]
        column_headers = data_CR.keys().tolist()
        CR_columns_to_transfer = [column_headers.index(CR_column_names[0]), column_headers.index(CR_column_names[1]), column_headers.index(CR_column_names[2]), column_headers.index(CR_column_names[3]), column_headers.index(CR_column_names[4]), column_headers.index(CR_column_names[5])]
        Headers_CR = data_CR.columns[CR_columns_to_transfer] 

        self.JF_column_names = ["Part Number","CURRENT Inventory","CURR. Orders","KBM Cons.","CURRENT Demand", "WIP"]
        column_headers = data_JF.keys().tolist()
        self.JF_columns_to_transfer = [column_headers.index(self.JF_column_names[0]), column_headers.index(self.JF_column_names[1]), column_headers.index(self.JF_column_names[2]), column_headers.index(self.JF_column_names[3]), column_headers.index(self.JF_column_names[4]), column_headers.index(self.JF_column_names[5])]
        Headers_JF = data_JF.columns[self.JF_columns_to_transfer] # part number, current inventory, current orders, KBM const, current demand, WIP

        ## Here we are making a list of lists of all the part data. Part numbers NOT included for this one. Part numbers are stored in part_numbers_existing_in_JF and part_number_ChoiceReport
        CR = np.array([data_CR[Headers_CR[1]].tolist(), data_CR[Headers_CR[2]].tolist(), data_CR[Headers_CR[3]].tolist(), data_CR[Headers_CR[4]].tolist(), data_CR[Headers_CR[5]].tolist()])
        JF = np.array([data_JF[Headers_JF[1]].tolist(), data_JF[Headers_JF[2]].tolist(), data_JF[Headers_JF[3]].tolist(), data_JF[Headers_JF[4]].tolist(), data_JF[Headers_JF[5]].tolist()])

        ## This is making the part numbers array from Jonesboro Forecast, ensuring no whitespace.
        part_numbers_existing_in_JF = []
        for cell in data_JF[Headers_JF[0]].tolist():
            if not pd.isnull(cell):
                cell = cell.strip()
            part_numbers_existing_in_JF.append(cell)

        ## Now we begin the iterating process through the whole excel sheet. Each part will be read from the Choice Report and searched for in the Jonesboro Forecast.
        self.parts_changed = []
        self.parts_not_found = []
        is_changed = False
        test = 0
        #We dont do the same part number array building in the Choice Report, so I strip each part number in the for loop as they're read to ensure no white space.
        part_number_ChoiceReport = data_CR[Headers_CR[0]].tolist()

        for ind, part_number in enumerate(part_number_ChoiceReport):
            
            part_number = part_number.strip()

            if part_number in part_numbers_existing_in_JF:
                is_changed = False
                temp = []
                part_number_ind = part_numbers_existing_in_JF.index(part_number)
                temp.append((part_number, part_number_ind+1))
                
                for i in range(5):
                    cell_CR_null = CR[i][ind]
                    cell_JF_null = JF[i][part_number_ind]

                    try:
                        cell_CR = int(cell_CR_null)
                    except:
                        cell_CR = 0

                    try:
                        cell_JF = int(float(cell_JF_null))
                    except:
                        cell_JF = 0

                    if cell_CR != cell_JF:
                        is_changed = True
                        temp.append((i, cell_JF, cell_CR))
                        


                if is_changed:
                    self.parts_changed.append(temp)

            else: # Part not found! Important to acknowledge.
                self.parts_not_found.append((part_number, ind))

        return







        # workbook.save(self.choiceReport.split(".xls")[0] + "_" + str(date) + ".xlsx")

        # workbook.close()

    def write_to_Jonesboro_Forecast(self):
        self.JF_columns_to_transfer.pop(0)

        workbook = openpyxl.load_workbook(self.jonesboroForecast_path)
        sheet = workbook.active

        test = 0
        for part in self.parts_changed:
            test += 1
            part_number = part[0]

            for change_in_quantity in part[1:]:
                Row = part_number[1] + 1
                Col = self.JF_columns_to_transfer[change_in_quantity[0]] + 1
                cell = sheet.cell(row = Row, column = Col)

                cell.value = change_in_quantity[2]

        workbook.save("done.xlsx")
        workbook.close()


    def print_output(self):


        built_string = "Parts changed:\n"
        # Adding text based on changed values
        for changed_part in self.parts_changed:
            part_num = changed_part.pop(0)
            
            built_string += "\n" + str(part_num[0]) + ": "
            number_for_spacing_output = 0
            for change in changed_part:
                built_string += str(self.JF_column_names[change[0] + 1]) + " changed from " + str(change[1]) + " to " + str(change[2]) + ". "
                number_for_spacing_output += 1
                if number_for_spacing_output == 2:
                    built_string += "\n"
            if number_for_spacing_output != 2:
                built_string += "\n"
        
        built_string += "\n\nParts not found:\n"

        for part in self.parts_not_found:
            built_string += f"\nDid not found part {part[0]} in the Jonesboro Forecast. This part was not updated. \nThis part was seen at position {str(part[1] + 2)} in the Choice Report.\n"
        # Create the main window
        root = tk.Tk()

        # Set the window title
        root.title("Results from data transfer")
        custom_font = font.Font(family="Arial", size=18)
        
        root.configure(bg="#ECECEC")

        # Create a Text widget for displaying text
        text_widget = tk.Text(root, font = custom_font)
        text_widget.pack()
        text_widget.insert(tk.END, built_string)

        # text_widget.insert(tk.END, "Neener!\n")
        def _quit():
            root.quit()
            root.destroy() 
        # Create a Close button to close the window
        close_button = tk.Button(root, text="Close", command=_quit)
        close_button.pack()

        # Run the main event loop for the results 
        root.mainloop()
        return    
    def run(self):

        while True:
            event, values = self.window.read()
            
            if event == sg.WIN_CLOSED or event == 'Exit':
                break
            
            elif event == 'Open Choice Report':
                self.choiceReport_path = sg.popup_get_file('Select a file', no_window=True)
                if self.choiceReport_path:
                    self.window['-OUTPUT-'].update(str(self.choiceReport_path))

            elif event == 'Open Jonesboro Forecast':
                self.jonesboroForecast_path = sg.popup_get_file('Select a file', no_window=True)
                if self.jonesboroForecast_path:
                    self.window['-OUTPUT-'].update(str(self.jonesboroForecast_path))
            
            elif event == 'Do Data transfer':
                if not self.choiceReport_path and not self.jonesboroForecast_path:
                    self.window['-OUTPUT-'].update('Need valid files.')
                    continue
                if not self.jonesboroForecast_path:
                    self.window['-OUTPUT-'].update('Need valid file for Jonesboro forecast.')
                    continue
                if not self.choiceReport_path:
                    self.window['-OUTPUT-'].update('Need valid file for Choice Report.')
                    continue

                self.read_data_files()
                self.write_to_Jonesboro_Forecast()
                self.window['-OUTPUT-'].update('Data transfer completed.')
                self.print_output()
                





if __name__ == "__main__":
    
    app = Application("Jonesboro Forecast Data Transfer")
    app.run()