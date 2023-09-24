import PySimpleGUI as sg
import pandas as pd
import openpyxl
import numpy as np
import tkinter as tk
from tkinter import font
import datetime

class Application:

    def __init__(self, appName, sourceFile, sourceFileColumns, targetFile, targetFileColumns, dateLocation):

        self.sourceFileName = sourceFile
        self.targetFileName = targetFile
        self.sourceFile_path = None
        self.targetFile_path = None
        self.save_file_name = None
        
        self.parts_changed = None
        self.parts_not_found = None

        self.sourceFileColumns = sourceFileColumns
        self.targetFileColumns = targetFileColumns

        self.dateLocation = dateLocation

        self.target_column_indices = None

        self.layout = [
            [sg.Button(f'Open {sourceFile}'), sg.Button(f'Open {targetFile}'), sg.Button('Do Data transfer')],
            [sg.Checkbox("Transfer WIP?", default=False, key="WIP_checkbox")],
            [sg.Text(size=(40, 1), key='-OUTPUT-')],
        ]
        self.window = sg.Window(appName, self.layout)

    # Reading excel file based on passed in path.
    def read_excel_file(self, file_name):
        try:
            data = pd.read_excel(file_name, sheet_name = 0)
            
        except:
            self.window['-OUTPUT-'].update(f'Invalid file for {file_name}.')
            return None
        return data
    
    # Getting all data to be transferred based on what was specified as input.
    def get_requested_data(self, data, columns_to_transfer):
        
        column_indices = []
        column_headers = data.keys().tolist()


            

        for column_name in columns_to_transfer:
            column_indices.append(column_headers.index(column_name))
        dataColumns = data.columns[column_indices]
        all_requested_data = data[dataColumns]

        return all_requested_data
    
    # Getting the indices for all relevant columns. This is important for storing data in output later.
    def get_column_indices(self, data, headers):
        columnNames = data.keys().tolist()
        columnIndices = []
        for header in headers:
            columnIndices.append(columnNames.index(header))
        return columnIndices
    
    # Getting all required data without part numbers.
    def get_data_except_part_numbers(self, data, columnNames):
        
        notNumpyArray = []
        indexArray = range(1, len(columnNames))
        for ind in indexArray:
            notNumpyArray.append(data[columnNames[ind]])
        data_without_partNumbers = np.array(notNumpyArray)
        return data_without_partNumbers
        
    # Getting the part numbers
    def get_part_numbers(self, data, columnNames):
        part_numbers = []
        
        for cell in data[columnNames[0]]:
            if not pd.isnull(cell):
                if type(cell) == str:
                    cell = cell.strip()
            part_numbers.append(cell)
        return part_numbers
    
    def get_save_file_name(self):
        folders = self.targetFile_path.split('/')
        return folders[-1]
        
    def read_data_files(self):
        #Getting all data from both source and target data files.
        # Source = file from which data is transferred. Target = file to which data is transferred.
        allSourceData = self.read_excel_file(self.sourceFile_path)
        allTargetData = self.read_excel_file(self.targetFile_path)     

        # Getting necessary data including part numbers
        requestedSourceData = self.get_requested_data(allSourceData, self.sourceFileColumns)
        requestedTargetData = self.get_requested_data(allTargetData, self.targetFileColumns)

        # Getting the column indices of the target file - important for storing data later in output file
        self.target_column_indices = self.get_column_indices(allTargetData, self.targetFileColumns)
        
        #Getting the data to be transferred without not including the part numbers.
        sourceDataNoPartNumbers = self.get_data_except_part_numbers(requestedSourceData, self.sourceFileColumns)
        targetDataNoPartNumbers = self.get_data_except_part_numbers(requestedTargetData, self.targetFileColumns)

        ## This is making the part numbers array from Jonesboro Forecast, ensuring no whitespace.
        sourcePartNumbers = self.get_part_numbers(requestedSourceData, self.sourceFileColumns)
        targetPartNumbers = self.get_part_numbers(requestedTargetData, self.targetFileColumns)

        ## Now we begin the iterating process through the whole excel sheet. Each part will be read 
        # from the source file and the target file, and their values will be compared. Differences 
        # between the values will be recorded..
        self.parts_changed = []
        self.parts_not_found = []
        is_changed = False

        for ind, part_number in enumerate(sourcePartNumbers):
            
            if part_number in targetPartNumbers:
                is_changed = False
                part_changes = []
                part_number_ind = targetPartNumbers.index(part_number)
                part_changes.append((part_number, part_number_ind+1))
                
                for i in range(len(self.targetFileColumns) - 1):
                    cell_source_null = sourceDataNoPartNumbers[i][ind]
                    target_cell_null = targetDataNoPartNumbers[i][part_number_ind]

                    try:
                        sourceCell = int(cell_source_null)
                    except:
                        sourceCell = 0

                    try:
                        targetCell = int(float(target_cell_null))
                    except:
                        targetCell = 0

                    if sourceCell != targetCell:
                        is_changed = True
                        part_changes.append((i, targetCell, sourceCell))
                        
                if is_changed:
                    self.parts_changed.append(part_changes)

            else: # Part not found! Important to acknowledge.
                self.parts_not_found.append((part_number, ind))

        return
    
    def write_date(self, sheet):
        currentDateCell = sheet.cell(row = self.dateLocation[0],column = self.dateLocation[1])
        currentDate = currentDateCell.value
        bothDates = currentDate.split('|')
        dateToStore = f"Last updated: {datetime.date.today()} | "

        if(self.do_WIP):
            dateToStore += f"WIP: {datetime.date.today()}"
        else:
            dateToStore += bothDates[1]

        currentDateCell.value = dateToStore
        return

    def write_to_output(self):

        workbook = openpyxl.load_workbook(self.targetFile_path)
        sheet = workbook.worksheets[0]

        for part in self.parts_changed:
            part_number = part[0]
            Row = part_number[1]

            for change_in_quantity in part[1:]:
                Col = self.target_column_indices[change_in_quantity[0] + 1]
                cell = sheet.cell(row = Row + 1, column = Col + 1)
                cell.value = change_in_quantity[2]

        self.write_date(sheet)

        workbook.save("results_" + self.save_file_name)
        workbook.close()

    def print_output(self):
        
        built_string = f"The following columns were transferred from {self.sourceFileName} to {self.targetFileName}:\n"

        for ind in range(len(self.targetFileColumns) - 1):
            ind += 1
            built_string += f"\"{self.sourceFileColumns[ind]}\" -> \"{self.targetFileColumns[ind]}\"\n"
        
        built_string += "\nSpecific changes:\n"
        # Adding text based on changed values
        for changed_part in self.parts_changed:
            part_num = changed_part.pop(0)
            built_string += "\n" + str(part_num[0]) + ": "
            number_for_spacing_output = 0
            for change in changed_part:
                built_string += str(self.targetFileColumns[change[0] + 1]) + " changed from " + str(change[1]) + " to " + str(change[2]) + ". "
                number_for_spacing_output += 1
                if number_for_spacing_output == 2:
                    built_string += "\n"
            if number_for_spacing_output != 2:
                built_string += "\n"
        
        built_string += "\n\nParts not found:\n"

        for part in self.parts_not_found:
            built_string += f"\nDid not found part {part[0]} in {self.targetFileName}. This part was not updated. \nThis part was seen at row {str(part[1] + 2)} in {self.sourceFileName}\n"
        # Create the main window
        root = tk.Tk()

        # Set the window title
        root.title("Results from data transfer")
        custom_font = font.Font(family="Arial", size=18)
        
        root.configure(bg="#ECECEC")

        # Create a Text widget for displaying text
        text_widget = tk.Text(root, font = custom_font)
        text_widget.pack()
        text_widget.insert(tk.END, built_string)

        # text_widget.insert(tk.END, "Neener!\n")
        def _quit():
            root.quit()
            root.destroy() 
        # Create a Close button to close the window
        close_button = tk.Button(root, text="Close", command=_quit)
        close_button.pack()

        # Run the main event loop for the results 
        root.mainloop()
        return    
    
    def run(self):

        while True:
            event, values = self.window.read()
            
            if event == sg.WIN_CLOSED or event == 'Exit':
                break
            
            elif event == f'Open {self.sourceFileName}':
                self.sourceFile_path = sg.popup_get_file('Select a file', no_window=True)
                if self.sourceFile_path:
                    self.window['-OUTPUT-'].update(str(self.sourceFile_path))

            elif event == f'Open {self.targetFileName}':
                self.targetFile_path = sg.popup_get_file('Select a file', no_window=True)
                if self.targetFile_path:
                    self.window['-OUTPUT-'].update(str(self.targetFile_path))
            
            elif event == 'Do Data transfer':
                if not self.sourceFile_path and not self.targetFile_path:
                    self.window['-OUTPUT-'].update('Need valid files.')
                    continue
                if not self.sourceFile_path:
                    self.window['-OUTPUT-'].update(f'Need valid file for {self.sourceFileName}.')
                    continue
                if not self.targetFile_path:
                    self.window['-OUTPUT-'].update(f'Need valid file for {self.targetFileName}.')
                    continue
                self.do_WIP = values['WIP_checkbox']
                self.save_file_name = self.get_save_file_name()
                if not self.do_WIP:
                    self.targetFileColumns.pop(-1)
                    self.sourceFileColumns.pop(-1)
                self.read_data_files()
                self.write_to_output()
                self.window['-OUTPUT-'].update('Data transfer completed.')
                self.print_output()
                

                
if __name__ == "__main__":

    app = Application(appName= "VS Stocking Summary Data Transfer", sourceFile= 'VS Report', sourceFileColumns= ["ICPROD", "Total On Hand","Consigned", "Required", "SumOfWIP"], targetFile= 'Stocking Summary', targetFileColumns=["T&B     Part #", "Total Inv:  JBS & T&B", "VS Cons  On Hand", "Current Demand", "WIP"], dateLocation = (1,1))
    app.run()